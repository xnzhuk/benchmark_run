import logging
import os
import time
import psutil
import subprocess

import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Font

def firestrike_data():
    # 判断firestrike_input.xml文件是否存在
    if os.path.exists("C:\\3DMark_result\\firestrike_input.xml"):
        # 解析XML文件
        tree = ET.parse('C:\\3DMark_result\\firestrike_input.xml')
        root = tree.getroot()

        # 提取数字并构建字典
        scores = {}
        prefixes = {
            'firestrikeoverallscorepforpass': 'overall',
            'firestrikephysicsscorepforpass': 'physics',
            'firestrikegraphicsscorepforpass': 'graphics',
            'firestrikecombinedscorepforpass': 'combined',
            'firestrikegt1p': 'gt1p',
            'firestrikegt2p': 'gt2p',
            'firestrikephysicsp': 'physicsp',
            'firestrikecombinedp': 'combinedp'
        }
        for element in root.iter():
            tag = element.tag
            if tag in prefixes:
                prefix = prefixes[tag]
                score = element.text
                scores[prefix] = score

        # 写入TXT文件
        with open('.\\3DMark\\firestrike_output.txt', 'w') as file:
            for prefix, score in scores.items():l
                line = f"{prefix}: {score}\n"
                file.write(line)

        # 加载Excel文件并选择工作表
        workbook = load_workbook("Performance_Record_Empty.xlsx")

        worksheet_name = "3DMark"
        if worksheet_name in workbook.sheetnames:
            sheet = workbook[worksheet_name]
            print(f"成功打开工作表 '{worksheet_name}'")

            # 获取当前电源状态
            def get_power_state():
                battery = psutil.sensors_battery()
                if battery.power_plugged:
                    return "AC"  # 电源已连接
                else:
                    return "DC"  # 电源未连接

            # 根据电源状态执行不同的操作
            def process_data_based_on_power_state(power_state):
                if power_state == "AC":
                    # AC 模式下的处理逻辑
                    cell_values = {
                        "firestrike_overall_ac": 'overall',
                        "firestrike_physics_ac": 'physics',
                        "firestrike_graphics_ac": 'graphics',
                        "firestrike_combined_ac": 'combined',
                        "firestrike_gt1p_ac": 'gt1p',
                        "firestrike_gt2p_ac": 'gt2p',
                        "firestrike_physicsp_ac": 'physicsp',
                        "firestrike_combinedp_ac": 'combinedp'
                    }
                    used_identifiers = set()

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value in cell_values:
                                cell_location = cell.coordinate
                                column_index = cell.column + 1  # 下一列

                                identifier = cell_values[cell.value]
                                if identifier not in used_identifiers:
                                    used_identifiers.add(identifier)
                                    if cell.offset(0, 1).value is None:  # 检查当前单元格的下一列是否为空
                                        sheet.cell(row=cell.row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到单元格")
                                    else:
                                        empty_row = cell.row + 1
                                        while sheet.cell(row=empty_row, column=column_index).value is not None:
                                            empty_row += 1
                                        sheet.cell(row=empty_row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到下一个空行的单元格")

                                    break

                elif power_state == "DC":
                    # DC 模式下的处理逻辑
                    cell_values = {
                        "firestrike_overall_dc": 'overall',
                        "firestrike_physics_dc": 'physics',
                        "firestrike_graphics_dc": 'graphics',
                        "firestrike_combined_dc": 'combined',
                        "firestrike_gt1p_dc": 'gt1p',
                        "firestrike_gt2p_dc": 'gt2p',
                        "firestrike_physicsp_dc": 'physicsp',
                        "firestrike_combinedp_dc": 'combinedp'
                    }
                    used_identifiers = set()

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value in cell_values:
                                cell_location = cell.coordinate
                                column_index = cell.column + 1  # 下一列

                                identifier = cell_values[cell.value]
                                if identifier not in used_identifiers:
                                    used_identifiers.add(identifier)
                                    if cell.offset(0, 1).value is None:  # 检查当前单元格的下一列是否为空
                                        sheet.cell(row=cell.row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到单元格")
                                    else:
                                        empty_row = cell.row + 1
                                        while sheet.cell(row=empty_row, column=column_index).value is not None:
                                            empty_row += 1
                                        sheet.cell(row=empty_row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到下一个空行的单元格")

                                    break

                else:
                    # 未知电源状态
                    print("无法识别当前电源状态！")

            # 获取当前电源状态
            power_state = get_power_state()
            print("当前电源状态：", power_state)

            # 根据电源状态执行相应的操作
            process_data_based_on_power_state(power_state)

            # 创建一个11号微软雅黑字体对象
            font = Font(name='微软雅黑', size=10)

            # 遍历所有单元格，并设置字体
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
            print(f"已将数据写入到单元格的同一行递增列中")
        else:
            print(f"未找到工作表 '{worksheet_name}'")

        # 保存修改后的Excel文件
        workbook.save("Performance_Record_Empty.xlsx")
        print("已保存修改后的Excel文件")
    else:
        print("C:\\3DMark_result\\firestrike_input.xml文件不存在")

def nightraid_data():
    # 判断nightraid_input.xml文件是否存在
    if os.path.exists("C:\\3DMark_result\\nightraid_input.xml"):

        # 解析XML文件
        tree = ET.parse('C:\\3DMark_result\\nightraid_input.xml')
        root = tree.getroot()

        # 提取数字并构建字典
        scores = {}
        prefixes = {
            'NightRaidPerformance3DMarkScoreForPass': '3DMarkScore',
            'NightRaidPerformanceGraphicsScoreForPass': 'GraphicsScore',
            'NightRaidPerformanceCPUScoreForPass': 'CPUScore',
            'NightRaidPerformanceGraphicsTest1': 'GraphicsTest1',
            'NightRaidPerformanceGraphicsTest2': 'GraphicsTest2',
            'NightRaidCpuP': 'CpuP',
        }
        for element in root.iter():
            tag = element.tag
            if tag in prefixes:
                prefix = prefixes[tag]
                score = element.text
                scores[prefix] = score

        # 写入TXT文件
        with open('.\\3DMark\\nightraid_output.txt', 'w') as file:
            for prefix, score in scores.items():
                line = f"{prefix}: {score}\n"
                file.write(line)

        # 加载Excel文件并选择工作表
        workbook = load_workbook("Performance_Record_Empty.xlsx")

        worksheet_name = "3DMark"
        if worksheet_name in workbook.sheetnames:
            sheet = workbook[worksheet_name]
            print(f"成功打开工作表 '{worksheet_name}'")

            # 获取当前电源状态
            def get_power_state():
                battery = psutil.sensors_battery()
                if battery.power_plugged:
                    return "AC"  # 电源已连接
                else:
                    return "DC"  # 电源未连接

            # 根据电源状态执行不同的操作
            def process_data_based_on_power_state(power_state):
                if power_state == "AC":
                    # AC 模式下的处理逻辑
                    cell_values = {
                        "nightraid_3DMarkScore_ac": '3DMarkScore',
                        "nightraid_GraphicsScore_ac": 'GraphicsScore',
                        "nightraid_CPUScore_ac": 'CPUScore',
                        "nightraid_GraphicsTest1_ac": 'GraphicsTest1',
                        "nightraid_GraphicsTest2_ac": 'GraphicsTest2',
                        "nightraid_CpuP_ac": 'CpuP',
                    }
                    used_identifiers = set()

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value in cell_values:
                                cell_location = cell.coordinate
                                column_index = cell.column + 1  # 下一列

                                identifier = cell_values[cell.value]
                                if identifier not in used_identifiers:
                                    used_identifiers.add(identifier)
                                    if cell.offset(0, 1).value is None:  # 检查当前单元格的下一列是否为空
                                        sheet.cell(row=cell.row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到单元格")
                                    else:
                                        empty_row = cell.row + 1
                                        while sheet.cell(row=empty_row, column=column_index).value is not None:
                                            empty_row += 1
                                        sheet.cell(row=empty_row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到下一个空行的单元格")

                                    break

                elif power_state == "DC":
                    # DC 模式下的处理逻辑
                    cell_values = {
                        "nightraid_3DMarkScore_ac": '3DMarkScore',
                        "nightraid_GraphicsScore_ac": 'GraphicsScore',
                        "nightraid_CPUScore_ac": 'CPUScore',
                        "nightraid_GraphicsTest1_ac": 'GraphicsTest1',
                        "nightraid_GraphicsTest2_ac": 'GraphicsTest2',
                        "nightraid_CpuP_ac": 'CpuP',
                    }
                    used_identifiers = set()

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value in cell_values:
                                cell_location = cell.coordinate
                                column_index = cell.column + 1  # 下一列

                                identifier = cell_values[cell.value]
                                if identifier not in used_identifiers:
                                    used_identifiers.add(identifier)
                                    if cell.offset(0, 1).value is None:  # 检查当前单元格的下一列是否为空
                                        sheet.cell(row=cell.row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到单元格")
                                    else:
                                        empty_row = cell.row + 1
                                        while sheet.cell(row=empty_row, column=column_index).value is not None:
                                            empty_row += 1
                                        sheet.cell(row=empty_row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到下一个空行的单元格")

                                    break

                else:
                    # 未知电源状态
                    print("无法识别当前电源状态！")

            # 获取当前电源状态
            power_state = get_power_state()
            print("当前电源状态：", power_state)

            # 根据电源状态执行相应的操作
            process_data_based_on_power_state(power_state)

            # 创建一个11号微软雅黑字体对象
            font = Font(name='微软雅黑', size=10)

            # 遍历所有单元格，并设置字体
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
            print(f"已将数据写入到单元格的同一行递增列中")
        else:
            print(f"未找到工作表 '{worksheet_name}'")

        # 保存修改后的Excel文件
        workbook.save("Performance_Record_Empty.xlsx")
        print("已保存修改后的Excel文件")
    else:
        print("C:\\3DMark_result\\nightraid_input.xml文件不存在")

def timespy_data():
    # 判断timespy_input.xml文件是否存在
    if os.path.exists("C:\\3DMark_result\\timespy_input.xml"):
        # 解析XML文件
        tree = ET.parse('C:\\3DMark_result\\timespy_input.xml')
        root = tree.getroot()

        # 提取数字并构建字典
        scores = {}
        prefixes = {
            'TimeSpyPerformanceCPUScoreForPass': 'CPUScore',
            'TimeSpyPerformance3DMarkScoreForPass': '3DMarkScore',
            'TimeSpyPerformanceGraphicsScoreForPass': 'GraphicsScore',
            'TimeSpyPerformanceGraphicsTest1': 'GraphicsTest1',
            'TimeSpyPerformanceGraphicsTest2': 'GraphicsTest2',
            'TimeSpyPerformanceCpuSection2': 'CpuSection2',
        }
        for element in root.iter():
            tag = element.tag
            if tag in prefixes:
                prefix = prefixes[tag]
                score = element.text
                scores[prefix] = score

        # 写入TXT文件
        with open('.\\3DMark\\timespy_output.txt', 'w') as file:
            for prefix, score in scores.items():
                line = f"{prefix}: {score}\n"
                file.write(line)

        # 加载Excel文件并选择工作表
        workbook = load_workbook("Performance_Record_Empty.xlsx")

        worksheet_name = "3DMark"
        if worksheet_name in workbook.sheetnames:
            sheet = workbook[worksheet_name]
            print(f"成功打开工作表 '{worksheet_name}'")

            # 获取当前电源状态
            def get_power_state():
                battery = psutil.sensors_battery()
                if battery.power_plugged:
                    return "AC"  # 电源已连接
                else:
                    return "DC"  # 电源未连接

            # 根据电源状态执行不同的操作
            def process_data_based_on_power_state(power_state):
                if power_state == "AC":
                    # AC 模式下的处理逻辑
                    cell_values = {
                        "timespy_CPUScore_ac": 'CPUScore',
                        "timespy_3DMarkScore_ac": '3DMarkScore',
                        "timespy_GraphicsScore_ac": 'GraphicsScore',
                        "timespy_GraphicsTest1_ac": 'GraphicsTest1',
                        "timespy_GraphicsTest2_ac": 'GraphicsTest2',
                        "timespy_CpuSection2_ac": 'CpuSection2',
                    }
                    used_identifiers = set()

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value in cell_values:
                                cell_location = cell.coordinate
                                column_index = cell.column + 1  # 下一列

                                identifier = cell_values[cell.value]
                                if identifier not in used_identifiers:
                                    used_identifiers.add(identifier)
                                    if cell.offset(0, 1).value is None:  # 检查当前单元格的下一列是否为空
                                        sheet.cell(row=cell.row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到单元格")
                                    else:
                                        empty_row = cell.row + 1
                                        while sheet.cell(row=empty_row, column=column_index).value is not None:
                                            empty_row += 1
                                        sheet.cell(row=empty_row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到下一个空行的单元格")

                                    break

                elif power_state == "DC":
                    # DC 模式下的处理逻辑
                    cell_values = {
                        "timespy_CPUScore_ac": 'CPUScore',
                        "timespy_3DMarkScore_ac": '3DMarkScore',
                        "timespy_GraphicsScore_ac": 'GraphicsScore',
                        "timespy_GraphicsTest1_ac": 'GraphicsTest1',
                        "timespy_GraphicsTest2_ac": 'GraphicsTest2',
                        "timespy_CpuSection2_ac": 'CpuSection2',
                    }
                    used_identifiers = set()

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value in cell_values:
                                cell_location = cell.coordinate
                                column_index = cell.column + 1  # 下一列

                                identifier = cell_values[cell.value]
                                if identifier not in used_identifiers:
                                    used_identifiers.add(identifier)
                                    if cell.offset(0, 1).value is None:  # 检查当前单元格的下一列是否为空
                                        sheet.cell(row=cell.row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到单元格")
                                    else:
                                        empty_row = cell.row + 1
                                        while sheet.cell(row=empty_row, column=column_index).value is not None:
                                            empty_row += 1
                                        sheet.cell(row=empty_row, column=column_index, value=float(scores[identifier]))

                                        print(f"找到单元格 '{cell.value}'，位置为 '{cell_location}'")
                                        print(f"已填写值 '{scores[identifier]}' 到下一个空行的单元格")

                                    break

                else:
                    # 未知电源状态
                    print("无法识别当前电源状态！")

            # 获取当前电源状态
            power_state = get_power_state()
            print("当前电源状态：", power_state)

            # 根据电源状态执行相应的操作
            process_data_based_on_power_state(power_state)

            # 创建一个11号微软雅黑字体对象
            font = Font(name='微软雅黑', size=10)

            # 遍历所有单元格，并设置字体
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
            print(f"已将数据写入到单元格的同一行递增列中")
        else:
            print(f"未找到工作表 '{worksheet_name}'")

        # 保存修改后的Excel文件
        workbook.save("Performance_Record_Empty.xlsx")
        print("已保存修改后的Excel文件")
    else:
        print("C:\\3DMark_result\\timespy_input.xml文件不存在")


def main():
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    option = get_user_input("请选择操作：1. firestrike；2. nightraid：3.timespy")

    if option == 1:
        option_one()
    elif option == 2:
        option_two()
    elif option == 3:
        option_three()
    else:
        print("无效的选项。")

def get_user_input(prompt):
    while True:
        try:
            return int(input(prompt))
        except ValueError:
            print("请输入一个有效的整数。")

def option_one():
    num_rounds = int(input("请输入圈数："))
    interval = int(input("请输入每圈间隔时间（秒）："))

    for _ in range(num_rounds):

        # 在指定路径下以管理员权限执行批处理语句
        process = subprocess.Popen('start /wait /b cmd /c 3DMark.bat', shell=True)
        # 等待脚本执行完全结束
        process.wait()
        print('3DMark安装')

        # 在C:\路径下新建3DMark_result文件夹
        folder_path = r"C:\3DMark_result"
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        print('在C:\路径下新建3DMark_result文件夹')

        # 在指定路径下以管理员权限执行批处理语句
        process = subprocess.Popen('"C:\\Program Files\\UL\\3DMark\\3DMarkCmd.exe" -register=3DM-TICFT-20230927-2KZF6-XD5KT-XZS3K-RKMEH', shell=True)
        # 等待脚本执行完全结束
        process.wait()
        print('3DMark激活')

        # 在指定路径下以管理员权限执行批处理语句
        process = subprocess.Popen('"C:\\Program Files\\UL\\3DMark\\3DMarkCmd.exe" --definition=firestrike.3dmdef --loop=1 --export=C:\\3DMark_result\\firestrike_input.xml --online=off --systeminfo=off --audio=off', shell=True)
        # 等待脚本执行完全结束
        process.wait()
        print('Run test')

        time.sleep(interval)

        firestrike_data()

def option_two():
    num_rounds = int(input("请输入圈数："))
    interval = int(input("请输入每圈间隔时间（秒）："))

    for _ in range(num_rounds):

        # 在指定路径下以管理员权限执行批处理语句
        process = subprocess.Popen('start /wait /b cmd /c 3DMark.bat', shell=True)
        # 等待脚本执行完全结束
        process.wait()
        print('3DMark安装')

        # 在C:\路径下新建3DMark_result文件夹
        folder_path = r"C:\3DMark_result"
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        print('在C:\路径下新建3DMark_result文件夹')

        # 在指定路径下以管理员权限执行批处理语句
        process = subprocess.Popen('"C:\\Program Files\\UL\\3DMark\\3DMarkCmd.exe" -register=3DM-TICFT-20230927-2KZF6-XD5KT-XZS3K-RKMEH', shell=True)
        # 等待脚本执行完全结束
        process.wait()
        print('3DMark激活')

        # 在指定路径下以管理员权限执行批处理语句
        process = subprocess.Popen('"C:\\Program Files\\UL\\3DMark\\3DMarkCmd.exe" --definition=nightraid.3dmdef --loop=1 --export=C:\\3DMark_result\\nightraid_input.xml --online=off --systeminfo=off --audio=off', shell=True)
        # 等待脚本执行完全结束
        process.wait()
        print('Run test')

        time.sleep(interval)

        nightraid_data()

def option_three():
    num_rounds = int(input("请输入圈数："))
    interval = int(input("请输入每圈间隔时间（秒）："))

    for _ in range(num_rounds):

        # 在指定路径下以管理员权限执行批处理语句
        process = subprocess.Popen('start /wait /b cmd /c 3DMark.bat', shell=True)
        # 等待脚本执行完全结束
        process.wait()
        print('3DMark安装')

        # 在C:\路径下新建3DMark_result文件夹
        folder_path = r"C:\3DMark_result"
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        print('在C:\路径下新建3DMark_result文件夹')

        # 在指定路径下以管理员权限执行批处理语句
        process = subprocess.Popen('"C:\\Program Files\\UL\\3DMark\\3DMarkCmd.exe" -register=3DM-TICFT-20230927-2KZF6-XD5KT-XZS3K-RKMEH', shell=True)
        # 等待脚本执行完全结束
        process.wait()
        print('3DMark激活')

        # 在指定路径下以管理员权限执行批处理语句
        process = subprocess.Popen('"C:\\Program Files\\UL\\3DMark\\3DMarkCmd.exe" --definition=timespy.3dmdef --loop=1 --export=C:\\3DMark_result\\nightraid_input.xml --online=off --systeminfo=off --audio=off', shell=True)
        # 等待脚本执行完全结束
        process.wait()
        print('Run test')

        time.sleep(interval)

        timespy_data()

if __name__ == "__main__":
    main()