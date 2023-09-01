import logging
import os
import re
import shutil
import subprocess
import time

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def process_data():
    # 判断single_result.txt文件是否存在
    if os.path.exists("Cinebenchr23\\single_result.txt"):
        # 读取single_result.txt文件内容
        with open("Cinebenchr23\\single_result.txt", "r") as file:
            content = file.read()

        # 使用正则表达式提取大括号中间的所有数字
        matches = re.findall(r"Values:\s*\{(.+?)\}", content)
        numbers = []
        for match in matches:
            numbers.extend(re.findall(r"\d+\.\d{2}", match))

        # 将提取到的数字写入single_output.txt文件
        with open("Cinebenchr23\\single_output.txt", "w") as output_file:
            output_file.write("\n".join(numbers))

        # # 打开同目录下的Performance_Record_Empty.xlsx表格文件
        # current_dir = os.path.dirname(os.path.abspath(__file__))
        # file_path = os.path.join(current_dir, "Performance_Record_Empty.xlsx")

        # 加载Excel文件并选择工作表
        workbook = load_workbook("Performance_Record_Empty.xlsx")

        # worksheet = workbook.active
        worksheet_name = "Cinebenchr23"
        if worksheet_name in workbook.sheetnames:
            sheet = workbook[worksheet_name]
            print(f"成功打开工作表 '{worksheet_name}'")

            # 搜索单元格
            cell_value = "CPU(Single Core)"
            cell_location = None

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == cell_value:
                        cell_location = cell.coordinate
                        break
                if cell_location:
                    break

            if cell_location:
                print(f"找到单元格 '{cell_value}'，位置为 '{cell_location}'")

                # 获取列号和起始行号
                column_index = cell.column + 1  # 下一列
                start_row = cell.row

                # 逐个写入数字到同一列的递增行
                for i, number in enumerate(numbers):
                    row = start_row + i
                    column = get_column_letter(column_index)
                    sheet[f"{column}{row}"].value = float(number)

                # 计算平均值
                avg_row = start_row
                avg_column = get_column_letter(column_index + 1)  # 下一列
                avg_formula = f"=AVERAGE({column}{start_row}:{column}{row})"
                sheet[f"{avg_column}{avg_row}"].value = avg_formula

                # 创建一个11号微软雅黑字体对象
                font = Font(name='微软雅黑', size=10)

                # 遍历所有单元格，并设置字体
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.font = font

                print(f"已将数字写入到单元格的同一列递增行中")
            else:
                print(f"未找到单元格 '{cell_value}'")
        else:
            print(f"未找到工作表 '{worksheet_name}'")

        # 保存修改后的Excel文件
        workbook.save("Performance_Record_Empty.xlsx")
        print("已保存修改后的Excel文件")
    else:
        print("Cinebenchr23\\single_result.txt文件不存在")

    # 判断multi_result.txt文件是否存在
    if os.path.exists("Cinebenchr23\\multi_result.txt"):
        # 读取multi_result.txt文件内容
        with open("Cinebenchr23\\multi_result.txt", "r") as file:
            content = file.read()

        # 使用正则表达式提取大括号中间的所有数字
        matches = re.findall(r"Values:\s*\{(.+?)\}", content)
        numbers = []
        for match in matches:
            numbers.extend(re.findall(r"\d+\.\d{2}", match))

        # 将提取到的数字写入multi_output.txt文件
        with open("Cinebenchr23\\multi_output.txt", "w") as output_file:
            output_file.write("\n".join(numbers))

        # # 打开同目录下的Performance_Record_Empty.xlsx表格文件
        # current_dir = os.path.dirname(os.path.abspath(__file__))
        # file_path = os.path.join(current_dir, "Performance_Record_Empty.xlsx")

        # 加载Excel文件并选择工作表
        workbook = load_workbook("Performance_Record_Empty.xlsx")

        # worksheet = workbook.active
        worksheet_name = "Cinebenchr23"
        if worksheet_name in workbook.sheetnames:
            sheet = workbook[worksheet_name]
            print(f"成功打开工作表 '{worksheet_name}'")

            # 搜索单元格
            cell_value = "CPU(Multi Core)"
            cell_location = None

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == cell_value:
                        cell_location = cell.coordinate
                        break
                if cell_location:
                    break

            if cell_location:
                print(f"找到单元格 '{cell_value}'，位置为 '{cell_location}'")

                # 获取列号和起始行号
                column_index = cell.column + 1  # 下一列
                start_row = cell.row

                # 逐个写入数字到同一列的递增行
                for i, number in enumerate(numbers):
                    row = start_row + i
                    column = get_column_letter(column_index)
                    sheet[f"{column}{row}"].value = float(number)

                # 计算平均值
                avg_row = start_row
                avg_column = get_column_letter(column_index + 1)  # 下一列
                avg_formula = f"=AVERAGE({column}{start_row}:{column}{row})"
                sheet[f"{avg_column}{avg_row}"].value = avg_formula

                # 创建一个11号微软雅黑字体对象
                font = Font(name='微软雅黑', size=10)

                # 遍历所有单元格，并设置字体
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.font = font

                print(f"已将数字写入到单元格的同一列递增行中")
            else:
                print(f"未找到单元格 '{cell_value}'")
        else:
            print(f"未找到工作表 '{worksheet_name}'")

        # 保存修改后的Excel文件
        workbook.save("Performance_Record_Empty.xlsx")
        print("已保存修改后的Excel文件")
    else:
        print("multi_result.txt文件不存在")

def main():
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    option = get_user_input("请选择操作：1. single；2. multi：")

    if option == 1:
        option_one()
    elif option == 2:
        option_two()
    else:
        print("无效的选项。")

def get_user_input(prompt):
    while True:
        try:
            return int(input(prompt))
        except ValueError:
            print("请输入一个有效的整数。")

def option_one():
    num_rounds = get_user_input("请输入圈数：")
    interval = get_user_input("请输入每圈间隔时间（秒）：")

    def delete_files(folder):
        for root, dirs, files in os.walk(folder):
            for file in files:
                file_path = os.path.join(root, file)
                os.remove(file_path)
                print("Deleted file:", file_path)
            for dir in dirs:
                dir_path = os.path.join(root, dir)
                shutil.rmtree(dir_path)
                print("Deleted folder:", dir_path)
    folder_path = os.path.expanduser("~") + "\\AppData\\Roaming\\MAXON\\"
    delete_files(folder_path)

    for _ in range(num_rounds):
        command = 'start /b /wait cmd.exe /C ".\\Cinebenchr23\\Cinebench.exe g_CinebenchCpuXTest=false g_CinebenchCpu1Test=true g_CinebenchMinimumTestDuration=1 >> .\\Cinebenchr23\\single_result.txt"'
        subprocess.call(command, shell=True)
        print("Single被执行了")
        time.sleep(interval)
    process_data()

def option_two():
    num_rounds = get_user_input("请输入圈数：")
    interval = get_user_input("请输入每圈间隔时间（秒）：")

    def delete_files(folder):
        for root, dirs, files in os.walk(folder):
            for file in files:
                file_path = os.path.join(root, file)
                os.remove(file_path)
                print("Deleted file:", file_path)
            for dir in dirs:
                dir_path = os.path.join(root, dir)
                shutil.rmtree(dir_path)
                print("Deleted folder:", dir_path)
    folder_path = os.path.expanduser("~") + "\\AppData\\Roaming\\MAXON\\"
    delete_files(folder_path)

    for _ in range(num_rounds):
        command = 'start /b /wait cmd.exe /C ".\\Cinebenchr23\\Cinebench.exe g_CinebenchCpuXTest=true g_CinebenchCpu1Test=false g_CinebenchMinimumTestDuration=1 >> .\\Cinebenchr23\\multi_result.txt"'
        subprocess.call(command, shell=True)
        print("Multi被执行了")
        time.sleep(interval)
    process_data()

if __name__ == "__main__":
    main()